import time
import os
import json
import pickle
from pathlib import Path
from datetime import datetime, timedelta

# Import libraries for web scraping
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager # Make sure you're using the correct manager
import io
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import openpyxl

# Import Google Drive libraries
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2 import service_account
from googleapiclient.discovery import build
import json
import os

# Configuration constants
URLS = {
    'main': 'https://www.dsebd.org/mkt_depth_3.php',
    'prices': 'https://dsebd.org/latest_share_price_scroll_by_change.php'
}
XPATHS = {
    'ticker': ('//*[@id="RightBody"]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[2]/a'),
    'ltp': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[3]'),
    'high': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[4]'),
    'low': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[5]'),
    'close': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[6]'),
    'ycp': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[7]'),
    'change': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[8]'),
    'trade': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[9]'),
    'value': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[10]'),
    'volume': ('/html/body/div[2]/section/div/div[3]/div[1]/div[2]/div[1]/div[2]/table/tbody', '/tr/td[11]'),
    'search_box': (By.NAME, "inst"),
    'buy_price_1': '//*[@id="RightBody"]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[1]/table/tbody/tr[3]/td[1]/div',
    'buy_volume_1': '//*[@id="RightBody"]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[1]/table/tbody/tr[3]/td[2]/div',
    'sell_price_1': '//*[@id="RightBody"]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/table/tbody/tr[3]/td[1]/div',
    'sell_volume_1': '//*[@id="RightBody"]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/table/tbody/tr[3]/td[2]/div',
    'buy_price_2': '//*[@id="RightBody"]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[1]/table/tbody/tr[4]/td[1]/div',
    'buy_volume_2': '//*[@id="RightBody"]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[1]/table/tbody/tr[4]/td[2]/div',
    'sell_price_2': '//*[@id="RightBody"]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/table/tbody/tr[4]/td[1]/div',
    'sell_volume_2': '//*[@id="RightBody"]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/table/tbody/tr[4]/td[2]/div'
}

# Google Drive setup
SCOPES = ['https://www.googleapis.com/auth/drive.file']
CREDENTIALS_FILE = 'credentials.json'
TOKEN_FILE = 'token.json'
DRIVE_FOLDER_ID = os.environ.get('DRIVE_FOLDER_ID')




def get_drive_service():
    """Authenticate and create Drive service using a service account"""
    try:
        # Get service account credentials from environment variable
        service_account_info = json.loads(os.environ.get('SERVICE_ACCOUNT_KEY'))
        print(f"Service account keys present: {', '.join(service_account_info.keys())}")
        
        # Create credentials from service account
        creds = service_account.Credentials.from_service_account_info(
            service_account_info, scopes=SCOPES)
        
        # Build and return the Drive service
        service = build('drive', 'v3', credentials=creds)
        print("Drive service created successfully using service account")
        return service
        
    except Exception as e:
        print(f"Error with Drive authentication: {str(e)}")
        raise






def initialize_driver():
    """Create and configure Chrome driver instance"""
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--window-size=1920x1080')

    print("Setting up Chrome Service using webdriver-manager...")
    try:
        # Get driver path
        driver_path = ChromeDriverManager().install()
        print(f"--- DEBUG: Initial path: {driver_path} ---")

        # Fix path if needed
        if 'THIRD_PARTY_NOTICES' in driver_path:
            corrected_path = os.path.join(os.path.dirname(driver_path), 'chromedriver')
            print(f"--- DEBUG: Corrected path: {corrected_path} ---")
            driver_path = corrected_path

        # Verify file exists
        if not os.path.isfile(driver_path):
            raise FileNotFoundError(f"Chromedriver not found at: {driver_path}")

        # Set execute permissions (crucial fix)
        os.chmod(driver_path, 0o755)  # This is the critical permission fix
        print(f"--- DEBUG: Set permissions for: {driver_path} ---")

        # Initialize service
        service = ChromeService(executable_path=driver_path)
        print("Chrome Service setup complete.")

        # Create driver
        driver = webdriver.Chrome(service=service, options=chrome_options)
        return driver

    except Exception as e:
        print(f"WebDriver init error: {str(e)}")
        raise





def get_bd_time():
    """Get current Bangladesh time (UTC+6)"""
    return datetime.utcnow() + timedelta(hours=6)





def setup_workbook():
    """Initialize or load Excel workbook"""
    date_str = get_bd_time().strftime('%d.%m.%Y')
    filename = f'Market_Depth_Auto_{date_str}.xlsx'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "market depth"
    headers = [
        "Date", "Time", "Ticker",
        "Buy Price_1", "Buy Volume_1",
        "Buy Price_2", "Buy Volume_2",
        "Sell Price_1", "Sell Volume_1",
        "Sell Price_2", "Sell Volume_2",
        "LTP", "High", "Low", "Close",
        "YCP", "Price Change", "Trade", "Value (mn)", "Total Volume"
    ]
    sheet.append(headers)
    return workbook, sheet, filename





def upload_to_drive(filename, drive_service):
    """Upload or append data to an existing Google Drive file."""

    file_name = os.path.basename(filename)
    query = f"name='{file_name}' and '{DRIVE_FOLDER_ID}' in parents and trashed=false"

    results = drive_service.files().list(q=query).execute()
    items = results.get('files', [])

    if items:
        # File exists, append data
        file_id = items[0].get('id')
        print(f"File '{file_name}' found with ID: {file_id}. Appending data.")

        # Download existing file
        request = drive_service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()

        # Load existing workbook
        existing_workbook = openpyxl.load_workbook(io.BytesIO(fh.getvalue()))
        existing_sheet = existing_workbook.active

        # Load new data
        new_workbook = openpyxl.load_workbook(filename)
        new_sheet = new_workbook.active

        # Append new data to existing sheet
        for row in new_sheet.iter_rows(min_row=2):  # Start from row 2 (skip header)
            values = [cell.value for cell in row]
            existing_sheet.append(values)

        # Upload updated file
        buffer = io.BytesIO()
        existing_workbook.save(buffer)
        buffer.seek(0)

        media = MediaIoBaseUpload(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        updated_file = drive_service.files().update(fileId=file_id, media_body=media).execute()

        print(f"Data appended to '{file_name}' with ID: {file_id}")
        return file_id

    else:
        # File does not exist, create new file
        print(f"File '{file_name}' not found. Creating new file.")

        file_metadata = {
            'name': file_name,
            'parents': [DRIVE_FOLDER_ID]
        }

        media = MediaFileUpload(
            filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )

        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()

        print(f"File uploaded to Drive with ID: {file.get('id')}")
        return file.get('id')




def scrape_data(driver, driver_prices, sheet):
    """Main scraping logic"""
    bd_time = get_bd_time()
    date_str = bd_time.strftime("%x")
    time_str = bd_time.strftime("%X")

    try:
        driver_prices.get(URLS['prices'])
        time.sleep(7)  # Allow time for page load

        # Process top 10 tickers
        for i in range(1, 11):
            ticker_xpath = f"{XPATHS['ticker'][0]}[{i}]{XPATHS['ticker'][1]}"
            time.sleep(5)  # Allow time for page load
            try:
                ticker_element = driver_prices.find_element(By.XPATH, ticker_xpath)
                ticker = ticker_element.text

                # Initialize default values
                buy_price_1 = buy_volume_1 = 0
                buy_price_2 = buy_volume_2 = 0
                sell_price_1 = sell_volume_1 = 0
                sell_price_2 = sell_volume_2 = 0

                # Fetch market depth data
                driver.get(URLS['main'])
                search_box = driver.find_element(*XPATHS['search_box'])
                search_box.send_keys(ticker + Keys.RETURN)
                time.sleep(5)

                # Extract buy/sell prices and volumes with error handling
                try:
                    buy_price_1 = float(driver.find_element(By.XPATH, XPATHS['buy_price_1']).text)
                    buy_volume_1 = int(driver.find_element(By.XPATH, XPATHS['buy_volume_1']).text.replace(',', ''))
                except (NoSuchElementException, ValueError):
                    pass

                try:
                    buy_price_2 = float(driver.find_element(By.XPATH, XPATHS['buy_price_2']).text)
                    buy_volume_2 = int(driver.find_element(By.XPATH, XPATHS['buy_volume_2']).text.replace(',', ''))
                except (NoSuchElementException, ValueError):
                    pass

                try:
                    sell_price_1 = float(driver.find_element(By.XPATH, XPATHS['sell_price_1']).text)
                    sell_volume_1 = int(driver.find_element(By.XPATH, XPATHS['sell_volume_1']).text.replace(',', ''))
                except (NoSuchElementException, ValueError):
                    pass

                try:
                    sell_price_2 = float(driver.find_element(By.XPATH, XPATHS['sell_price_2']).text)
                    sell_volume_2 = int(driver.find_element(By.XPATH, XPATHS['sell_volume_2']).text.replace(',', ''))
                except (NoSuchElementException, ValueError):
                    pass

                # Append data to the sheet
                sheet.append([
                    date_str,
                    time_str,
                    ticker,
                    buy_price_1, buy_volume_1,
                    buy_price_2, buy_volume_2,
                    sell_price_1, sell_volume_1,
                    sell_price_2, sell_volume_2,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['ltp'][0]}[{i}]{XPATHS['ltp'][1]}").text,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['high'][0]}[{i}]{XPATHS['high'][1]}").text,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['low'][0]}[{i}]{XPATHS['low'][1]}").text,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['close'][0]}[{i}]{XPATHS['close'][1]}").text,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['ycp'][0]}[{i}]{XPATHS['ycp'][1]}").text,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['change'][0]}[{i}]{XPATHS['change'][1]}").text,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['trade'][0]}[{i}]{XPATHS['trade'][1]}").text,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['value'][0]}[{i}]{XPATHS['value'][1]}").text,
                    driver_prices.find_element(By.XPATH, f"{XPATHS['volume'][0]}[{i}]{XPATHS['volume'][1]}").text.replace(',', '')
                ])
                print(f"Saved data for {ticker} at {time_str}")

            except Exception as e:
                print(f"Error processing ticker {i}: {str(e)}")
                continue

    except Exception as e:
        print(f"Major error: {str(e)}")





def main():
    """Main execution function"""
    print("Starting market depth scraper...")
    driver = initialize_driver()
    driver_prices = initialize_driver()
    
    try:
        # Setup workbook
        workbook, sheet, filename = setup_workbook()
        
        # Scrape data
        scrape_data(driver, driver_prices, sheet)
        
        # Save file locally
        workbook.save(filename)
        print(f"Data saved to {filename}")
        
        # Upload to Google Drive
        drive_service = get_drive_service()
        upload_to_drive(filename, drive_service)
        
    except Exception as e:
        print(f"Error in main function: {str(e)}")
    
    finally:
        driver.quit()
        driver_prices.quit()

if __name__ == "__main__":
    main()
